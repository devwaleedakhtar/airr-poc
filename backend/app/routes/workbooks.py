from __future__ import annotations

import os
import tempfile
from typing import Any, Dict, List, Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from ..core.cloudinary import download_to_temp, upload_raw
from cloudinary.utils import private_download_url
from ..core.db import db_dependency
from ..core.config import settings
from ..repositories import workbooks_repo
from ..schemas.workbooks import (
    ConvertRequest,
    ConvertResponse,
    ExtractRequest,
    ExtractResponse,
    UploadWorkbookResponse,
)
from ..services.converter_service import convert_excel_sheet_to_pdf
from ..services.converter_service import _ensure_xlsx
from ..services import o365_converter_service


router = APIRouter(prefix="/workbooks", tags=["workbooks"])


@router.post("/upload", response_model=UploadWorkbookResponse)
async def upload_workbook(
    file: UploadFile = File(...),
    db=Depends(db_dependency),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    if not (file.filename.lower().endswith(".xlsx") or file.filename.lower().endswith(".xlsm") or file.filename.lower().endswith(".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xlsm, and .xls are supported")

    # Write to temp
    fd, tmp_path = tempfile.mkstemp(suffix=os.path.splitext(file.filename)[1])
    os.close(fd)
    converted_tmp: Optional[str] = None
    try:
        content = await file.read()
        with open(tmp_path, "wb") as f:
            f.write(content)

        # Discover sheets
        from openpyxl import load_workbook

        try:
            read_path = _ensure_xlsx(tmp_path)
            if read_path != tmp_path:
                converted_tmp = read_path
            wb = load_workbook(read_path, read_only=True, data_only=True)
            # Only expose visible sheets in the UI to match analyst expectations
            sheets = [ws.title for ws in wb.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read workbook: {e}")

        # Pre-generate id to use in Cloudinary path
        workbook_id = await workbooks_repo.generate_id()

        # Upload original to Cloudinary
        folder = f"airr-poc/workbooks/{workbook_id}"
        upload_res = upload_raw(tmp_path, public_id="original", folder=folder)
        original_url = upload_res.get("secure_url") or upload_res.get("url") or ""
        original_public_id = upload_res.get("public_id")
        original_format = upload_res.get("format")

        # Persist workbook record
        await workbooks_repo.create(
            db,
            {
                "_id": ObjectId(workbook_id),
                "filename": file.filename,
                "content_type": file.content_type,
                "sheets": sheets,
                "original_url": original_url,
                "original_public_id": original_public_id,
                "original_format": original_format,
            },
        )

        return UploadWorkbookResponse(workbook_id=workbook_id, sheets=sheets)
    finally:
        if converted_tmp and os.path.exists(converted_tmp):
            try:
                os.remove(converted_tmp)
            except FileNotFoundError:
                pass
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except FileNotFoundError:
                pass


@router.post("/{workbook_id}/convert", response_model=ConvertResponse)
async def convert_workbook(
    workbook_id: str,
    payload: ConvertRequest,
    db=Depends(db_dependency),
):
    doc = await workbooks_repo.get(db, workbook_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Workbook not found")
    if payload.sheet_name not in doc.get("sheets", []):
        raise HTTPException(status_code=400, detail="Invalid sheet name")

    tmp_xlsx: Optional[str] = None
    pdf_path: Optional[str] = None
    try:
        # Download original (used for LibreOffice path; Office 365 path streams via Graph)
        original_public_id = doc.get("original_public_id")
        original_format = doc.get("original_format") or os.path.splitext(doc.get("filename", ""))[1].lstrip(".")
        if original_public_id:
            signed = private_download_url(
                original_public_id,
                original_format or "",
                resource_type="raw",
                type="private",
            )
            tmp_xlsx = download_to_temp(signed, suffix=("." + (original_format or "xlsx")))
        else:
            tmp_xlsx = download_to_temp(doc["original_url"], suffix=os.path.splitext(doc.get("filename", ""))[1])

        # Convert single sheet to PDF using configured backend
        try:
            backend = (settings.converter_backend or "libreoffice").lower()
            result = None

            if backend in ("office365", "o365", "graph", "hybrid"):
                try:
                    result = o365_converter_service.convert_via_office365(
                        doc,
                        workbook_id,
                        payload.sheet_name,
                    )
                except Exception:
                    if backend == "hybrid" and tmp_xlsx:
                        result = convert_excel_sheet_to_pdf(tmp_xlsx, payload.sheet_name)
                    else:
                        raise
            else:
                result = convert_excel_sheet_to_pdf(tmp_xlsx, payload.sheet_name)

            pdf_path = result.pdf_path

            graph_item_id = getattr(result, "graph_item_id", None)
            if graph_item_id:
                await workbooks_repo.set_graph_item_id(db, workbook_id, graph_item_id)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Conversion failed: {e}")

        # Upload PDF to Cloudinary
        from ..services.converter_service import _slugify  # reuse

        sheet_slug = _slugify(payload.sheet_name)
        folder = f"airr-poc/pdfs/{workbook_id}"
        upload_res = upload_raw(pdf_path, public_id=sheet_slug, folder=folder)
        pdf_url = upload_res.get("secure_url") or upload_res.get("url") or ""
        pdf_public_id = upload_res.get("public_id")
        pdf_format = upload_res.get("format") or "pdf"

        await workbooks_repo.set_pdf_for_sheet(
            db,
            workbook_id,
            payload.sheet_name,
            pdf_url,
            public_id=pdf_public_id,
            fmt=pdf_format,
        )

        return ConvertResponse(pdf_url=pdf_url)
    finally:
        if pdf_path and os.path.exists(pdf_path):
            try:
                os.remove(pdf_path)
            except FileNotFoundError:
                pass
        if tmp_xlsx and os.path.exists(tmp_xlsx):
            try:
                os.remove(tmp_xlsx)
            except FileNotFoundError:
                pass


@router.post("/{workbook_id}/extract", response_model=ExtractResponse)
async def extract_workbook(
    workbook_id: str,
    payload: ExtractRequest,
    db=Depends(db_dependency),
):
    from ..services.extractor_service import extract_from_pdf

    doc = await workbooks_repo.get(db, workbook_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Workbook not found")

    # Determine PDF URL
    sheet_name = payload.sheet_name
    pdfs = doc.get("pdfs", {})
    pdf_public_ids = doc.get("pdf_public_ids", {})
    pdf_formats = doc.get("pdf_formats", {})
    pdf_url: Optional[str] = None
    if sheet_name and sheet_name in pdfs:
        pdf_url = pdfs[sheet_name]
    elif not sheet_name and isinstance(pdfs, dict) and pdfs:
        # pick any existing converted sheet
        sheet_name = list(pdfs.keys())[0]
        pdf_url = pdfs[sheet_name]
    
    if not pdf_url:
        raise HTTPException(status_code=400, detail="No PDF found for extraction. Convert a sheet first or provide sheet_name.")

    tmp_pdf: Optional[str] = None
    try:
        # Download PDF
        if pdf_public_ids.get(sheet_name):
            signed = private_download_url(
                pdf_public_ids[sheet_name],
                pdf_formats.get(sheet_name) or "pdf",
                resource_type="raw",
                type="private",
            )
            tmp_pdf = download_to_temp(signed, suffix=".pdf")
        else:
            tmp_pdf = download_to_temp(pdf_url, suffix=".pdf")

        # Extract
        try:
            result = extract_from_pdf(tmp_pdf)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Extraction failed: {e}")

        # Persist session
        from ..repositories import sessions_repo

        session_id = await sessions_repo.create(
            db,
            {
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "pdf_url": pdf_url,
                "extracted_json": result.extracted_json,
                "confidences": result.confidences,
                "inferred_tables": result.inferred_tables,
                "warnings": result.warnings,
                "text_snippets": result.text_snippets,
            },
        )

        return ExtractResponse(
            session_id=session_id,
            extracted_json=result.extracted_json,
            confidences=result.confidences,
            inferred_tables=result.inferred_tables,
            warnings=result.warnings,
        )
    finally:
        if tmp_pdf and os.path.exists(tmp_pdf):
            try:
                os.remove(tmp_pdf)
            except FileNotFoundError:
                pass
