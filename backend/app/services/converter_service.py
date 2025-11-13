from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from typing import Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from ..core.config import settings


@dataclass
class ConversionResult:
    pdf_path: str


def _slugify(text: str) -> str:
    return "".join(c.lower() if c.isalnum() else "-" for c in text).strip("-")


def _compute_used_range(ws) -> Tuple[int, int, int, int]:
    min_row, min_col = None, None
    max_row, max_col = 0, 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                r, c = cell.row, cell.column
                min_row = r if min_row is None else min(min_row, r)
                min_col = c if min_col is None else min(min_col, c)
                max_row = max(max_row, r)
                max_col = max(max_col, c)
    if min_row is None or min_col is None:
        # default to A1
        return 1, 1, 1, 1
    return min_row, min_col, max_row, max_col


def _autofit_columns(ws) -> None:
    col_max = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            col_max[cell.column_letter] = max(col_max.get(cell.column_letter, 0), len(val))
    for col_letter, max_len in col_max.items():
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

def _convert_xls_to_xlsx(xls_path: str) -> str:
    """Convert legacy .xls to .xlsx using LibreOffice and return a temp .xlsx path."""
    out_dir = tempfile.mkdtemp()
    soffice = _find_soffice_binary()
    cmd = [
        soffice,
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        out_dir,
        xls_path,
    ]
    subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    base = os.path.splitext(os.path.basename(xls_path))[0]
    candidate = os.path.join(out_dir, f"{base}.xlsx")
    if not os.path.exists(candidate):
        for name in os.listdir(out_dir):
            if name.lower().endswith(".xlsx"):
                candidate = os.path.join(out_dir, name)
                break
    if not os.path.exists(candidate):
        raise RuntimeError("XLS to XLSX conversion failed: no XLSX produced")
    fd, final_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    shutil.move(candidate, final_path)
    return final_path

def _ensure_xlsx(src_path: str) -> str:
    """Ensure the returned path is .xlsx/.xlsm; convert .xls if needed for openpyxl."""
    ext = os.path.splitext(src_path)[1].lower()
    if ext == ".xls":
        return _convert_xls_to_xlsx(src_path)
    return src_path


def _prepare_single_sheet_workbook(src_path: str, sheet_name: str) -> str:
    """Prepare a temp workbook preserving formatting and print layout.

    - Opens the original workbook (preserving VBA if present).
    - Hides all non-target sheets and sets the target active.
    - Applies print/fit settings and sets print area to used range.
    - Saves one temp copy (xlsm if source was xlsm, else xlsx).
    """
    ext = os.path.splitext(src_path)[1].lower()
    keep_vba = ext == ".xlsm"

    processed_src = _ensure_xlsx(src_path)
    wb = load_workbook(filename=processed_src, keep_vba=keep_vba, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

    ws = wb[sheet_name]

    # Fit to page width and compute print area from used range
    _autofit_columns(ws)
    min_row, min_col, max_row, max_col = _compute_used_range(ws)
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    ws.print_area = f"{start}:{end}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    if not ws.sheet_properties.pageSetUpPr:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    # Hide all other sheets; keep only target visible and active
    for name in wb.sheetnames:
        wb[name].sheet_state = "visible" if name == sheet_name else "hidden"
    wb.active = wb.sheetnames.index(sheet_name)

    suffix = ".xlsm" if keep_vba else ".xlsx"
    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    wb.save(tmp_path)
    return tmp_path


def _find_soffice_binary() -> str:
    if settings.libreoffice_path and os.path.exists(settings.libreoffice_path):
        return settings.libreoffice_path
    which = shutil.which("soffice")
    if which:
        return which
    raise RuntimeError(
        "LibreOffice 'soffice' not found. Install LibreOffice and set LIBREOFFICE_PATH or add 'soffice' to PATH."
    )


def _convert_xlsx_to_pdf(xlsx_path: str) -> str:
    out_dir = tempfile.mkdtemp()
    soffice = _find_soffice_binary()
    cmd = [
        soffice,
        "--headless",
        "--convert-to",
        "pdf:calc_pdf_Export",
        "--outdir",
        out_dir,
        xlsx_path,
    ]
    subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    # find generated pdf
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    pdf_candidate = os.path.join(out_dir, f"{base}.pdf")
    if not os.path.exists(pdf_candidate):
        # fallback: first pdf in out_dir
        for name in os.listdir(out_dir):
            if name.lower().endswith(".pdf"):
                pdf_candidate = os.path.join(out_dir, name)
                break
    if not os.path.exists(pdf_candidate):
        raise RuntimeError("PDF conversion failed: no PDF produced")
    # move to a stable temp file
    fd, final_path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    shutil.move(pdf_candidate, final_path)
    return final_path


def convert_excel_sheet_to_pdf(src_xlsx_path: str, sheet_name: str) -> ConversionResult:
    single_sheet_path = _prepare_single_sheet_workbook(src_xlsx_path, sheet_name)
    pdf_path = _convert_xlsx_to_pdf(single_sheet_path)
    return ConversionResult(pdf_path=pdf_path)
