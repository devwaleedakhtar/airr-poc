from __future__ import annotations

import shutil
import tempfile
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Tuple

import cloudinary.utils
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel

from ..core.cloudinary import upload_raw
from ..core.config import settings
from ..schemas.mapping import MappingResult
from ..schemas.sessions import ExportAppliedField, ExportResponse

TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "constants" / "model_input_sheet.xlsx"
TARGET_SHEET_NAME = "PERPETUITY MIDDLEMAN INPUT TAB"

# Canonical scalar fields -> cell references in the template sheet.
SCALAR_CELL_MAP: Dict[Tuple[str, str], str] = {
    # Growth assumptions
    ("growth_assumptions", "market_rent_growth"): "I35",
    ("growth_assumptions", "affordable_rent_growth"): "I36",
    ("growth_assumptions", "other_income_growth"): "I37",
    ("growth_assumptions", "controllables_growth"): "I38",
    ("growth_assumptions", "taxes_growth"): "I39",
    ("growth_assumptions", "insurance_growth"): "I40",
    # Project timeline
    ("project_timeline", "land_closing_date"): "C11",
    ("project_timeline", "construction_start_month"): "C13",
    ("project_timeline", "first_units_delivered_month"): "C16",
    # Revenue and lease-up
    ("revenue_and_leaseup", "vacancy_pct"): "C22",
    ("revenue_and_leaseup", "loss_to_lease_pct"): "C23",
    ("revenue_and_leaseup", "bad_debt_pct"): "C24",
    ("revenue_and_leaseup", "model_units"): "C25",
    ("revenue_and_leaseup", "concessions_lease_up_months"): "C27",
    ("revenue_and_leaseup", "leased_units_per_month"): "C36",
    ("revenue_and_leaseup", "renewal_probability_pct"): "C37",
    ("revenue_and_leaseup", "lease_term_months"): "C38",
    # Operating expenses
    ("operating_expenses", "payroll"): "C48",
    ("operating_expenses", "utilities"): "C49",
    ("operating_expenses", "turnover"): "C50",
    ("operating_expenses", "contract_services"): "C51",
    ("operating_expenses", "repairs_maintenance"): "C52",
    ("operating_expenses", "leasing_marketing"): "C53",
    ("operating_expenses", "general_admin"): "C54",
    ("operating_expenses", "other_expenses"): "C55",
    ("operating_expenses", "management_fee_pct"): "C58",
    ("operating_expenses", "insurance"): "C59",
    ("operating_expenses", "property_taxes"): "C60",
    ("operating_expenses", "other_taxes_fees"): "C61",
    ("operating_expenses", "replacement_reserves"): "C64",
    # Senior loan terms
    ("senior_loan_terms", "loan_to_cost_pct"): "F5",
    ("senior_loan_terms", "interest_type"): "F6",
    ("senior_loan_terms", "curve"): "F7",
    ("senior_loan_terms", "sofr_spread_pct"): "F8",
    ("senior_loan_terms", "sofr_floor_pct"): "F9",
    ("senior_loan_terms", "sofr_cap_pct"): "F10",
    ("senior_loan_terms", "interest_only_period_months"): "F11",
    ("senior_loan_terms", "amortization_schedule_years"): "F12",
    ("senior_loan_terms", "initial_term_months"): "F13",
    ("senior_loan_terms", "origination_fee_pct"): "F15",
    ("senior_loan_terms", "rate_stepdown_dscr_multiple"): "F16",
    ("senior_loan_terms", "rate_stepdown_dy_pct"): "F17",
    ("senior_loan_terms", "stepdown_rate_pct"): "F18",
    ("senior_loan_terms", "exit_fee_pct"): "F20",
    # Preferred equity terms
    ("preferred_equity_terms", "has_preferred_equity"): "C93",
    ("preferred_equity_terms", "loan_to_cost_pct"): "C94",
    ("preferred_equity_terms", "initial_term_months"): "C95",
    ("preferred_equity_terms", "interest_type"): "C96",
    ("preferred_equity_terms", "sofr_spread_pct"): "C97",
    ("preferred_equity_terms", "sofr_floor_pct"): "C98",
    ("preferred_equity_terms", "total_interest_rate_pct"): "C99",
    ("preferred_equity_terms", "minimum_multiple"): "C100",
    ("preferred_equity_terms", "current_pay_pct"): "C101",
    ("preferred_equity_terms", "accrual_pct"): "C102",
    # Exit assumptions
    ("exit_assumptions", "sale_month"): "F25",
    ("exit_assumptions", "noi_type"): "F26",
    ("exit_assumptions", "sale_costs_pct"): "F27",
    ("exit_assumptions", "exit_cap_rate_mf_pct"): "F28",
    ("exit_assumptions", "exit_cap_rate_retail_pct"): "F29",
    # Tax reassessment at exit
    ("tax_reassessment_at_exit", "reassess_at_sale"): "E33",
    ("tax_reassessment_at_exit", "property_tax_millage_rate_pct"): "F34",
    ("tax_reassessment_at_exit", "county_assessment_pct"): "F35",
    ("tax_reassessment_at_exit", "market_value_as_pct_of_sale_price"): "F36",
    # Sources and uses
    ("sources_and_uses", "land_acquisition_cost"): "J13",
    ("sources_and_uses", "hard_costs_total"): "J14",
    ("sources_and_uses", "soft_costs_total"): "J15",
    ("sources_and_uses", "financing_costs"): "J16",
    ("sources_and_uses", "operating_reserve"): "J17",
    ("sources_and_uses", "senior_interest_reserve"): "J18",
}

UNIT_MIX_START_ROW = 23
UNIT_MIX_MAX_ROWS = 6
UNIT_MIX_COLUMNS = {
    "unit_type": "H",
    "num_units": "I",
    "avg_sf": "J",
    "rent": "K",
    "original_label": "H",
}

OTHER_INCOME_START_ROW = 70
OTHER_INCOME_MAX_ROWS = 16
OTHER_INCOME_COLUMNS = {
    "item_name": "B",
    "num_units": "C",
    "amount_per_month": "D",
}

WATERFALL_START_ROW = 45
WATERFALL_MAX_ROWS = 5
WATERFALL_COLUMNS = {
    "tier_name": "H",
    "lp_split_pct": "I",
    "gp_split_pct": "J",
    "hurdle_irr_pct": "L",
    "moic_multiple": "M",
    "dollar_amount": "N",
}


def _coerce_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, (datetime, date)):
        return to_excel(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        # Percent strings like "6%" -> 0.06
        if text.endswith("%"):
            try:
                return float(text.rstrip("%").replace(",", "")) / 100.0
            except Exception:
                return text
        # Currency / numeric strings
        cleaned = text.replace(",", "").replace("$", "")
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            return float(cleaned)
        except Exception:
            pass
        # ISO date
        try:
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
            return to_excel(dt)
        except Exception:
            return text
    return value


def _is_formula(cell) -> bool:
    val = cell.value
    if getattr(cell, "data_type", None) == "f":
        return True
    return isinstance(val, str) and val.strip().startswith("=")


def _clear_table_rows(
    ws, start_row: int, max_rows: int, column_map: Dict[str, str], skip_columns: set[str] | None = None
) -> None:
    targets = {col for col in column_map.values()}
    if skip_columns:
        targets -= set(skip_columns)
    for idx in range(max_rows):
        row = start_row + idx
        for col in targets:
            ws[f"{col}{row}"].value = None


def _apply_scalar_values(ws, mapped: Dict[str, Any], applied: List[ExportAppliedField]) -> None:
    for (table, field), cell in SCALAR_CELL_MAP.items():
        table_val = mapped.get(table)
        if not isinstance(table_val, dict):
            continue
        raw = table_val.get(field)
        coerced = _coerce_value(raw)
        if coerced is None:
            continue
        cell_obj = ws[cell]
        if _is_formula(cell_obj):
            continue
        cell_obj.value = coerced
        applied.append(ExportAppliedField(table=table, field=field, cell=cell, value=raw))


def _apply_unit_mix(ws, mapped: Dict[str, Any], applied: List[ExportAppliedField]) -> None:
    rows = mapped.get("unit_mix")
    if not isinstance(rows, list) or not rows:
        return
    # Preserve template formulas (e.g., PSF) by not clearing derived columns.
    _clear_table_rows(ws, UNIT_MIX_START_ROW, UNIT_MIX_MAX_ROWS, UNIT_MIX_COLUMNS, skip_columns={"L"})
    for idx, row in enumerate(rows[:UNIT_MIX_MAX_ROWS]):
        if not isinstance(row, dict):
            continue
        target_row = UNIT_MIX_START_ROW + idx
        for field, col in UNIT_MIX_COLUMNS.items():
            if field not in row:
                continue
            raw = row.get(field)
            coerced = _coerce_value(raw)
            if coerced is None:
                continue
            cell = f"{col}{target_row}"
            cell_obj = ws[cell]
            if _is_formula(cell_obj):
                continue
            cell_obj.value = coerced
            applied.append(ExportAppliedField(table="unit_mix", field=field, cell=cell, value=raw))


def _apply_other_income(ws, mapped: Dict[str, Any], applied: List[ExportAppliedField]) -> None:
    rows = mapped.get("other_income")
    if not isinstance(rows, list) or not rows:
        return
    _clear_table_rows(ws, OTHER_INCOME_START_ROW, OTHER_INCOME_MAX_ROWS, OTHER_INCOME_COLUMNS)
    for idx, row in enumerate(rows[:OTHER_INCOME_MAX_ROWS]):
        if not isinstance(row, dict):
            continue
        target_row = OTHER_INCOME_START_ROW + idx
        item_name = row.get("item_name")
        if item_name not in (None, ""):
            cell = f"{OTHER_INCOME_COLUMNS['item_name']}{target_row}"
            text = str(item_name)
            cell_obj = ws[cell]
            if not _is_formula(cell_obj):
                cell_obj.value = text
                applied.append(ExportAppliedField(table="other_income", field="item_name", cell=cell, value=item_name))
        for field, col in OTHER_INCOME_COLUMNS.items():
            if field == "item_name":
                continue
            raw = row.get(field)
            coerced = _coerce_value(raw)
            if coerced is None:
                continue
            cell = f"{col}{target_row}"
            cell_obj = ws[cell]
            if _is_formula(cell_obj):
                continue
            cell_obj.value = coerced
            applied.append(ExportAppliedField(table="other_income", field=field, cell=cell, value=raw))


def _apply_waterfall(ws, mapped: Dict[str, Any], applied: List[ExportAppliedField]) -> None:
    rows = mapped.get("waterfall")
    if not isinstance(rows, list) or not rows:
        return
    _clear_table_rows(ws, WATERFALL_START_ROW, WATERFALL_MAX_ROWS, WATERFALL_COLUMNS)
    fields_to_write = [
        "tier_name",
        "lp_split_pct",
        "gp_split_pct",
        "hurdle_irr_pct",
        "moic_multiple",
        "dollar_amount",
    ]
    for idx, row in enumerate(rows[:WATERFALL_MAX_ROWS]):
        if not isinstance(row, dict):
            continue
        target_row = WATERFALL_START_ROW + idx
        for field in fields_to_write:
            col = WATERFALL_COLUMNS.get(field)
            if not col:
                continue
            raw = row.get(field)
            if field == "tier_name":
                if raw in (None, ""):
                    continue
                text = str(raw)
                cell_obj = ws[f"{col}{target_row}"]
                if not _is_formula(cell_obj):
                    cell_obj.value = text
                    applied.append(ExportAppliedField(table="waterfall", field=field, cell=f"{col}{target_row}", value=raw))
                continue
            coerced = _coerce_value(raw)
            if coerced is None:
                continue
            cell = f"{col}{target_row}"
            cell_obj = ws[cell]
            if _is_formula(cell_obj):
                continue
            cell_obj.value = coerced
            applied.append(ExportAppliedField(table="waterfall", field=field, cell=cell, value=raw))


def export_mapping(session_id: str, mapping: MappingResult) -> ExportResponse:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template workbook not found at {TEMPLATE_PATH}")

    tmp_dir = Path(tempfile.mkdtemp())
    tmp_path = tmp_dir / f"{session_id}-model.xlsx"
    shutil.copy(TEMPLATE_PATH, tmp_path)

    wb = load_workbook(tmp_path)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{TARGET_SHEET_NAME}' not found in template")
    ws = wb[TARGET_SHEET_NAME]

    applied_fields: List[ExportAppliedField] = []
    _apply_scalar_values(ws, mapping.mapped, applied_fields)
    _apply_unit_mix(ws, mapping.mapped, applied_fields)
    _apply_other_income(ws, mapping.mapped, applied_fields)
    _apply_waterfall(ws, mapping.mapped, applied_fields)

    wb.save(tmp_path)

    folder = f"{settings.cloudinary_base_folder}/exports/{session_id}".rstrip("/")
    upload_res = upload_raw(str(tmp_path), public_id="model-input", folder=folder)

    # Generate signed URL for private file access (valid for 4 hours)
    public_id = upload_res.get("public_id", "")
    signed_url = cloudinary.utils.private_download_url(
        public_id,
        "xlsx",
        resource_type="raw",
        type="private",
        expires_at=int((datetime.utcnow().timestamp()) + 14400),
    )

    return ExportResponse(download_url=signed_url, applied_fields=applied_fields)
